import pymysql
import openpyxl
import time

def createwb(wbname):
    wb=openpyxl.Workbook()
    wb.save(filename=wbname)
    print ("New Excel："+wbname+" created")

def savetoexcel(data,fields,sheetname,wbname):
    print("Dumping database to Excel...")
    wb=openpyxl.load_workbook(filename=wbname)

    sheet=wb.active
    sheet.title=sheetname

    field=1
    for field in range(1,len(fields)+1):   # 写入表头
        _=sheet.cell(row=1,column=field,value=str(fields[field-1]))

    row1=1
    col1=0
    for row1 in range(2,len(data)+2):  # 写入数据
        for col1 in range(1,len(data[row1-2])+1):
            _=sheet.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    wb.save(filename=wbname)
    print("Data saved")

def db_to_list(db):
    try:
        conn = pymysql.connect(host='116.62.190.97', port=3306, user='remote', passwd='ali545783Tx.*', db='sf_taobao',charset="utf8")
        cur = conn.cursor()
    except:
        sys.exit('\n错误：数据库连接失败')
    select = 'SELECT * FROM ' + db
    cur.execute(select)
    lst = []
    for row in cur :
        rowlst = []
        for cell in row:
            rowlst.append(cell)
        lst.append(rowlst)
    cur.close()
    conn.close()
    return lst


db = 'db_list'
data = db_to_list(db)

wbname = db + '_' + time.strftime("%Y-%m-%d", time.localtime()) + '.xlsx'
createwb(wbname)

fields = ['ID', 'pid', 'itemUrl', 'status', 'title', 'initialPrice', 'currentPrice', 'consultPrice', 'marketPrice', 'sellOff', 'start', 'end',
          'timeToStart', 'timeToEnd', 'viewerCount', 'bidCount', 'delayCount', 'applyCount', 'xmppVersion', 'buyRestrictions', 'supportLoans', 'supportOrgLoan']
savetoexcel(data, fields, wbname, wbname)
